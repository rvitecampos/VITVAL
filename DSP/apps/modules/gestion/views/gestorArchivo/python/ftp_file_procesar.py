# -*- encoding: utf-8 -*-

# /---------------------------\
# | Developed by Luis Remicio |
# \ ------------------------- /

import sys, os, ftputil, base64, informixdb, codecs
from datetime import datetime, date, time
from xlrd import open_workbook, cellname, xldate_as_tuple
import openpyxl
# from openpyxl import load_workbook

id_solicitud = 23

class Estructura:

    def __init__(self, orden, tamanio, nombre):
        self.orden = orden
        self.tamanio = tamanio
        self.nombre =nombre

    def __repr__(self):
        return repr((self.orden, self.tamanio, self.nombre))


class FTP_procesar:

    id_solicitud = 0
    rs = []
    cursor = ''

    pathTmp = '/sistemas/weburbano/public_html/tmp/gestor_archivos/'

    estado = 0

    ftp_host_origen = ''
    ftp_user_origen = ''
    ftp_pass_origen = ''
    ftp_path_origen = ''
    ftp_file = ''

    ftp_host_destino = ''
    ftp_user_destino = ''
    ftp_pass_destino = ''
    ftp_path_destino = ''

    def __init__(self, id_solicitud):
        self.id_solicitud = id_solicitud

        '''
        Parameters of connection - Informix
        '''
        os.environ['INFORMIXSERVER'] = 'ol_urbano'
        conn = informixdb.connect('scm30@ol_urbano', user='informix', password='mh_1C$_2sX')
        self.cursor = conn.cursor(rowformat = informixdb.ROW_AS_DICT)
        self.rs = self.getDataSolicitud()
        self.rs = self.rs[0]

        self.estado = int(self.rs['estado'])

        self.ftp_host_origen = self.rs['ftp_server']
        self.ftp_user_origen = self.rs['ftp_user']
        self.ftp_pass_origen = self.rs['ftp_clave']
        self.ftp_path_origen = self.rs['ftp_path']
        self.ftp_file = self.rs['ftp_file']

        self.ftp_host_destino = self.rs['ftp_server_d']
        self.ftp_user_destino = self.rs['ftp_user_d']
        self.ftp_pass_destino = self.rs['ftp_clave_d']
        self.ftp_path_destino = self.rs['ftp_path_txt']

        self.getFileTmp()
        self.validar_file()

        # self.makeTxt()

    def getDataSolicitud(self):
        query_string = "call scm_gestor_ftp_chk_file("+str(self.id_solicitud)+");"
        self.cursor.execute(query_string)
        rs = self.cursor.fetchall()
        return rs

    def getEstructuraFile(self):
        query_string = "call scm_gestor_ftp_chk_file_struc("+str(self.id_solicitud)+");"
        self.cursor.execute(query_string)
        rs = self.cursor.fetchall()
        return rs

    def getFileTmp(self):
        with ftputil.FTPHost(self.ftp_host_origen, self.ftp_user_origen, self.ftp_pass_origen) as ftp_host:
            download = ftp_host.download_if_newer(self.ftp_path_origen + self.ftp_file, self.pathTmp + self.ftp_file)

    def setFileTmp(self):
        fileNameTxt = self.ftp_file
        fileNameTxt = os.path.splitext(fileNameTxt)[0] + '.txt'
        with ftputil.FTPHost(self.ftp_host_destino, self.ftp_user_destino, self.ftp_pass_destino) as ftp_host:
            upload = ftp_host.upload_if_newer(self.pathTmp + fileNameTxt, self.ftp_path_destino + fileNameTxt)

    def valida_none(self, cadena):
        cadena = '' if (cadena == None) else cadena
        return cadena

    def validar_file(self):
        if self.estado == 1:
            self.setFileTmp()
        elif self.estado == 2:
            fileName = self.pathTmp + self.ftp_file
            extension = os.path.splitext(fileName)[1][1:]

            fileNameTxt = self.ftp_file
            fileNameTxt = self.pathTmp + os.path.splitext(fileNameTxt)[0] + '.txt'

            rs = self.getEstructuraFile()

            aEstructura = []
            for col in rs:
                aEstructura.append(Estructura(col['orden'], col['tamanio'], col['nombre']))

            aEstructura = sorted(aEstructura, key=lambda estructura: estructura.orden)

            if extension == 'xls':
                '''
                Other libraries for files .xls
                '''
                workbook = open_workbook(fileName)
                worksheet = workbook.sheet_by_index(0)

                txt = codecs.open(fileNameTxt, "w", encoding='latin-1')
                for row_index in range(worksheet.nrows):
                    if row_index != 0:
                        cadena = ''
                        for col in aEstructura:
                            cadena+= self.valida_none(worksheet.cell(row_index,int(col.orden - 1)).value).strip()[:col.tamanio].ljust(col.tamanio)
                        txt.write(cadena + "\n")
                txt.close()
            else:
                workbook = openpyxl.load_workbook(filename = fileName)
                worksheet = workbook.get_active_sheet()

                i = 0
                txt = codecs.open(fileNameTxt, "w", encoding='utf-8')
                for cell in worksheet.rows:
                    if i != 0:
                        cadena = ''
                        for col in aEstructura:
                            cadena+= self.valida_none(cell[int(col.orden - 1)].value).strip()[:col.tamanio].ljust(col.tamanio)
                        txt.write(cadena + "\n")
                    
                    i+=1
                txt.close()

            self.setFileTmp()

            os.remove(fileName)
            os.remove(fileNameTxt)

obj = FTP_procesar(id_solicitud)
